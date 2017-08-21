# coding:utf-8

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import json
from datetime import datetime
import datetime as dt
import calendar
from openpyxl import Workbook

inputDateStr = '2017-07-01' #指定爬蟲起始日期
rawRecordList = []
itAttendanceList = []
userList = []
addList = []
removeList = []


def holidayChk(year): #假日清單產生function
    day31List = [1, 3, 5, 7, 8, 10, 12]
    day30List = [4, 6, 9, 11]
    tmpList = []
    for i in range(1, 13, 1):
        if i in day31List:
            for j in range(1, 32, 1):
                date = dt.date(year, i, j)
                if dt.date.weekday(date) == 5 or dt.date.weekday(date) == 6:
                    date = str(date)
                    tmpList.append(date)
        if i in day30List:
            for j in range(1, 31, 1):
                date = dt.date(year, i, j)
                if dt.date.weekday(date) == 5 or dt.date.weekday(date) == 6:
                    date = str(date)
                    tmpList.append(date)
        if i == 2:
            for j in range(1, 29, 1):
                date = dt.date(year, i, j)
                if dt.date.weekday(date) == 5 or dt.date.weekday(date) == 6:
                    date = str(date)
                    tmpList.append(date)
            try:
                date229 = dt.date(year, 2, 29)
                if dt.date.weekday(date229) == 5 or dt.date.weekday(date229) == 6:
                    tmpList.append(str(date229))
            except:
                pass

    chk_a = False
    chk_b = False
    with open("/Users/yunhan/Desktop/holidayModification.txt", "r") as f:  # 獲得假日增減資料
        for line in f.readlines():
            while chk_a:
                if line.strip() == '':
                    chk_a = False
                    break
                tmpList.append(line.strip())
                break
            if line.strip() == "Add Holiday:":
                chk_a = True
            while chk_b:
                if line.strip() == '':
                    chk_a = False
                    break
                tmpList.remove(line.strip())
                break
            if line.strip() == "Remove Holiday:":
                chk_b = True

    return tmpList

if __name__ == "__main__":
    holidayList = holidayChk(int(inputDateStr.split('-')[0]))


    with open("/Users/yunhan/Desktop/it_userList.txt", "r") as f:   #獲得員工ID
        for line in f.readlines():
            userList.append(line.strip())
    username = input("LDAP Username:")
    password = input("Password:")
    driver_path = '/Users/yunhan/driver/chromedriver'
    web = webdriver.Chrome(driver_path)
    web.get('https://talk.chatchat365.com/api/v3/oauth/gitlab/login')
    web.find_element_by_id('username').send_keys(username)
    web.find_element_by_id('password').send_keys(password)
    web.find_element_by_id('password').send_keys(Keys.ENTER)
    time.sleep(3)
    web.get('https://talk.chatchat365.com/smartbackend/channels/it-signin')
    time.sleep(10)


    # 載入爬蟲起始日期之後的資料
    inputTimeStr = inputDateStr + ' 00:00:00'
    inputDate = datetime.strptime(inputDateStr, "%Y-%m-%d")
    inputTime = datetime.strptime(inputTimeStr, "%Y-%m-%d %H:%M:%S")

    while True:
        n = 0
        while n < 10:
            try:
                web.find_element_by_class_name('more-messages-text').find_element_by_tag_name('span').click()
                break
            except:
                time.sleep(1)
                n += 1
                continue
        origin = web.find_elements_by_class_name('separator__text')[0].text
        year = origin.split(u'年')[0]
        month = origin.split(u'年')[1].split(u'月')[0]
        if len(month) < 2:
           month = '0'+month

        day = origin.split(u'年')[1].split(u'月')[1].split(u'日')[0]
        dataDateStr = year + '-' + month + '-' + day
        dataDate = datetime.strptime(dataDateStr, "%Y-%m-%d")
        if dataDate <= inputDate:
            break

        time.sleep(1)
    time.sleep(5)


    #將資料爬回
    number = len(web.find_elements_by_class_name('post__content'))
    print(u'將資料爬回:' + str(number) + u'筆')

    for j in range(0, number, 1):
        data = web.find_elements_by_class_name('post__content')[j]
        id = data.find_element_by_class_name('user-popover').text.strip()
        if id in userList:
            rawPostDate = data.find_element_by_class_name("post__time").get_attribute("datetime").split('T')[0]
            rawPostTime = data.find_element_by_class_name("post__time").get_attribute("datetime").split('T')[1].split('.')[0]
            postTimeStr = rawPostDate + ' ' + rawPostTime
            postTimeUtc = datetime.strptime(postTimeStr, "%Y-%m-%d %H:%M:%S")
            timeStamp = calendar.timegm(postTimeUtc.utctimetuple())
            localPostTime = datetime.fromtimestamp(timeStamp)
            localPostTimeStr = datetime.fromtimestamp(timeStamp).strftime('%Y-%m-%d %H:%M:%S')

            if localPostTime >= inputTime:
                rawRecord_dict = {
                    "ID": "",
                    "date": "",
                    "postTime":"",
                    "postMessage": "",
                    "status": "1"
                }
                rawRecord_dict['ID'] = id
                rawRecord_dict['postTime'] = localPostTimeStr.split(" ")[1]
                rawRecord_dict['date'] = localPostTimeStr.split(" ")[0]
                postMessage = data.find_element_by_class_name('post-message__text').text.strip()
                postMessageLower = postMessage.lower()
                if (postMessageLower.__contains__('in') or postMessageLower.__contains__('out')) \
                        and len(postMessageLower) < 15:
                    rawRecord_dict['status'] = '0'
                rawRecord_dict['postMessage'] = postMessageLower
                print(rawRecord_dict)
                rawRecordList.append(rawRecord_dict)

            else:
                continue
    # 清洗資料
    dateSet = set()
    for rawRecord in rawRecordList:
        dateSet.add(rawRecord['date'])
    for date in dateSet:
        tmp0Record = []
        tmpIdSet = set()
        for record in rawRecordList:
            if record['date'] == date and record['status'] == '0':
                tmp0Record.append(record)
                tmpIdSet.add(record['ID'])
        for name in tmpIdSet:
            tmpRecord = []
            for record in tmp0Record:
                if record['ID'] == name:
                    tmpRecord.append(record)

            num = len(tmpRecord)
            attendanceStatus_dict = {
                'ID': '',
                'date': '',
                'signIn': '',
                'signOut': '',
                'workHours': '',
                'status': '',
            }
            if num >= 2 and tmpRecord[0]['postMessage'].__contains__('in') and tmpRecord[num - 1][
                'postMessage'].__contains__('out'):
                attendanceStatus_dict['ID'] = name
                attendanceStatus_dict['date'] = date
                attendanceStatus_dict['signIn'] = tmpRecord[0]['postTime']
                attendanceStatus_dict['signOut'] = tmpRecord[num - 1]['postTime']
                signInTimeHourCal = tmpRecord[0]['postTime'].split(':')[0]
                signInTimeMinCal = tmpRecord[0]['postTime'].split(':')[1]
                signInTimeCal = date + ' ' + signInTimeHourCal + ':' + signInTimeMinCal + ':00'
                signOutTimeHourCal = tmpRecord[num - 1]['postTime'].split(':')[0]
                signOutTimeMinCal = tmpRecord[num - 1]['postTime'].split(':')[1]
                signOutTimeCal = tmpRecord[num - 1][
                                     'date'] + ' ' + signOutTimeHourCal + ':' + signOutTimeMinCal + ':00'
                signInTime = datetime.strptime(signInTimeCal, "%Y-%m-%d %H:%M:%S")
                signOutTime = datetime.strptime(signOutTimeCal, "%Y-%m-%d %H:%M:%S")

                if date not in holidayList:
                    if signOutTime - signInTime >= dt.timedelta(hours=9):
                        attendanceStatus_dict['status'] = '0'
                        attendanceStatus_dict['workHours'] = "+" + str(
                            signOutTime - signInTime - dt.timedelta(hours=9))
                    else:
                        attendanceStatus_dict['status'] = '3'
                        attendanceStatus_dict['workHours'] = "-" + str(
                            dt.timedelta(hours=9) - (signOutTime - signInTime))
                if date in holidayList:
                    attendanceStatus_dict['status'] = '0'
                    attendanceStatus_dict['workHours'] = "+" + str(signOutTime - signInTime)
                if num > 2:
                    if attendanceStatus_dict['status'] == '0':
                        attendanceStatus_dict['status'] = '2'
                    else:
                        attendanceStatus_dict['status'] = attendanceStatus_dict['status'] + ',2'

                itAttendanceList.append(attendanceStatus_dict)

            if num == 1:  # 將4錯誤訊息輸出
                tmpRecord[0]['status'] = '4'
                itAttendanceList.append(tmpRecord[0])
                attendanceStatus_dict['ID'] = name
                attendanceStatus_dict['date'] = date
                attendanceStatus_dict['status'] = '4'
                attendanceStatus_dict['workHours'] = '-'

                if tmpRecord[0]['postMessage'].__contains__('in'):
                    attendanceStatus_dict['signIn'] = tmpRecord[0]['postTime']
                    attendanceStatus_dict['signOut'] = '-'
                else:
                    attendanceStatus_dict['signIn'] = '-'
                    attendanceStatus_dict['signOut'] = tmpRecord[0]['postTime']

                itAttendanceList.append(attendanceStatus_dict)

        if date not in holidayList:  # 將5錯誤訊息輸出
            for idChk in userList:
                if idChk not in tmpIdSet:
                    absence_dict = {
                        "ID": "",
                        "date": "",
                        "postTime": "-",
                        "postMessage": "-",
                        "status": "5"
                    }
                    absence_dict['ID'] = idChk
                    absence_dict['date'] = date
                    itAttendanceList.append(absence_dict)

    for i in range(0, len(rawRecordList), 1):  # 將1錯誤訊息輸出
        if rawRecordList[i]['status'] == '1':
            itAttendanceList.append(rawRecordList[i])

    for i in range(0, len(itAttendanceList), 1):  # 將2,3錯誤訊息輸出
        statusList = ['0', '1', '4', '5']
        if itAttendanceList[i]['status'] not in statusList:
            workHoursChk_dict = {
                "ID": "",
                "date": "",
                "postTime": "-",
                "postMessage": "-",
                "status": ""
            }
            workHoursChk_dict['ID'] = itAttendanceList[i]['ID']
            workHoursChk_dict['date'] = itAttendanceList[i]['date']
            workHoursChk_dict['status'] = itAttendanceList[i]['status']
            itAttendanceList.append(workHoursChk_dict)


    # 寫成Excel(日期分頁)
    attendanceXlsx = Workbook()
    sheet = attendanceXlsx.active
    attendanceXlsx.remove_sheet(sheet)

    rawRecordSheet = attendanceXlsx.create_sheet('rawRecord')
    rawRecordSheet.cell(row=1, column=1, value='date')
    rawRecordSheet.cell(row=1, column=2, value='ID')
    rawRecordSheet.cell(row=1, column=3, value='postTime')
    rawRecordSheet.cell(row=1, column=4, value='postMessage')
    rawRecordSheet.column_dimensions['A'].width = 15
    rawRecordSheet.column_dimensions['B'].width = 15
    rawRecordSheet.column_dimensions['C'].width = 15
    rawRecordSheet.column_dimensions['D'].width = 15
    rowNum = 2
    for raw in rawRecordList:
        rawRecordSheet.cell(row=rowNum, column=1, value=raw['date'])
        rawRecordSheet.cell(row=rowNum, column=2, value=raw['ID'])
        rawRecordSheet.cell(row=rowNum, column=3, value=raw['postTime'])
        rawRecordSheet.cell(row=rowNum, column=4, value=raw['postMessage'])
        rowNum += 1

    statusReportSheet = attendanceXlsx.create_sheet('statusReport')
    statusReportSheet.cell(row=1, column=1, value='status code')
    statusReportSheet.cell(row=1, column=2, value='description')
    statusReportSheet.cell(row=2, column=1, value='1')
    statusReportSheet.cell(row=2, column=2, value='incorrect postMessage records')
    statusReportSheet.cell(row=3, column=1, value='2')
    statusReportSheet.cell(row=3, column=2, value='more than 2 postMessage records in a day')
    statusReportSheet.cell(row=4, column=1, value='3')
    statusReportSheet.cell(row=4, column=2, value='Work less than 8.5 hours')
    statusReportSheet.cell(row=5, column=1, value='4')
    statusReportSheet.cell(row=5, column=2, value='only 1 postMessage record in a day')
    statusReportSheet.cell(row=6, column=1, value='5')
    statusReportSheet.cell(row=6, column=2, value='Absence')

    statusReportSheet.cell(row=8, column=1, value='postDate')
    statusReportSheet.cell(row=8, column=2, value='ID')
    statusReportSheet.cell(row=8, column=3, value='status code')
    statusReportSheet.cell(row=8, column=4, value='postTime')
    statusReportSheet.cell(row=8, column=5, value='postMessage')
    statusReportSheet.column_dimensions['A'].width = 15
    statusReportSheet.column_dimensions['B'].width = 15
    statusReportSheet.column_dimensions['C'].width = 15
    statusReportSheet.column_dimensions['D'].width = 15
    statusReportSheet.column_dimensions['E'].width = 15
    statusReportSheet.column_dimensions['F'].width = 15

    R = 9
    dateList = list(dateSet)
    list.sort(dateList)
    for date in dateList:
        sheet = attendanceXlsx.create_sheet(date)
        sheet['A1'] = 'Date'
        sheet['B1'] = 'ID'
        sheet['C1'] = 'Sign In'
        sheet['D1'] = 'Sign Out'
        sheet['E1'] = 'Work Hours'
        sheet['F1'] = 'Status'
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15

        r = 2
        for i in range(0, len(itAttendanceList), 1):
            if itAttendanceList[i]['date'] == date and len(itAttendanceList[i]) == 6:
                sheet.cell(row=r, column=1, value=date)
                sheet.cell(row=r, column=2, value=itAttendanceList[i]['ID'])
                sheet.cell(row=r, column=3, value=itAttendanceList[i]['signIn'])
                sheet.cell(row=r, column=4, value=itAttendanceList[i]['signOut'])
                sheet.cell(row=r, column=5, value=itAttendanceList[i]['workHours'])
                sheet.cell(row=r, column=6, value=itAttendanceList[i]['status'])
                r += 1
            if itAttendanceList[i]['date'] == date and len(itAttendanceList[i]) == 5:
                statusReportSheet.cell(row=R, column=1, value=itAttendanceList[i]['date'])
                statusReportSheet.cell(row=R, column=2, value=itAttendanceList[i]['ID'])
                statusReportSheet.cell(row=R, column=3, value=itAttendanceList[i]['status'])
                statusReportSheet.cell(row=R, column=4, value=itAttendanceList[i]['postTime'])
                statusReportSheet.cell(row=R, column=5, value=itAttendanceList[i]['postMessage'])
                R += 1

    attendanceXlsx.save('/Users/yunhan/Desktop/it_Attendance.xlsx')
    print('Done')
    web.quit()



    # 寫成jason
    # encodedjson = json.dumps(itAttendanceList, ensure_ascii=False)
    # with open("/Users/yunhan/Desktop/itAttendance.json", "w") as f:
    #     f.write(encodedjson)
    #     f.close()
    # web.quit()
