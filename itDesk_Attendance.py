# coding:utf-8

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import json
from datetime import datetime
import datetime as dt
import calendar
from openpyxl import Workbook
from IT_Attendance import holidayChk

if __name__ == "__main__":
    inputDateStr = '2017-07-01' #指定爬蟲起始日期
    rawRecordList = []
    itDeskAttendanceList = []
    errorList = []
    userListRotat = []
    userListAdmini = []
    holidayList = holidayChk(int(inputDateStr.split('-')[0]))
    chk_a = False
    chk_b = False

    with open("/Users/yunhan/Desktop/itDesk_userList.txt", "r") as f: #獲得員工ID及班別
        for line in f.readlines():
            while chk_a:
                if line.strip() == '':
                    chk_a = False
                    break
                userListRotat.append(line.strip())
                break
            if line.strip() == "rotation_staff:":
                chk_a = True
            while chk_b:
                if line.strip() == '':
                    chk_a = False
                    break
                userListAdmini.append(line.strip())
                break
            if line.strip() == "admini_staff:":
                chk_b = True
    username = input("LDAP Username:")
    password = input("Password:")
    driver_path = '/Users/yunhan/driver/chromedriver'
    web = webdriver.Chrome(driver_path)
    web.get('pleaseEnterUrlforLogin')
    web.find_element_by_id('username').send_keys(username)
    web.find_element_by_id('password').send_keys(password)
    web.find_element_by_id('password').send_keys(Keys.ENTER)
    time.sleep(3)
    web.get('linkToForum')
    time.sleep(10)


    # 載入爬蟲起始日期之後的資料
    inputTimeStr = inputDateStr + ' 00:00:00'
    inputDate = datetime.strptime(inputDateStr, "%Y-%m-%d")
    inputTime = datetime.strptime(inputTimeStr, "%Y-%m-%d %H:%M:%S")

    while True:
        n = 0
        while n < 10:
            try:
                web.find_element_by_class_name('more-messages-text').find_element_by_tag_name('span').click()
                break
            except:
                time.sleep(1)
                n += 1
                continue
        origin = web.find_elements_by_class_name('separator__text')[0].text
        year = origin.split(u'年')[0]
        month = origin.split(u'年')[1].split(u'月')[0]
        if len(month) < 2:
           month = '0'+month

        day = origin.split(u'年')[1].split(u'月')[1].split(u'日')[0]
        dataDateStr = year + '-' + month + '-' + day
        dataDate = datetime.strptime(dataDateStr, "%Y-%m-%d")
        if dataDate <= inputDate:
            break

        time.sleep(1)
    time.sleep(5)


    #將資料爬回
    number = len(web.find_elements_by_class_name('post__content'))
    print(u'將資料爬回:' + str(number) + u'筆')

    for j in range(0, number, 1):
        data = web.find_elements_by_class_name('post__content')[j]
        id = data.find_element_by_class_name('user-popover').text.strip()
        if id in userListAdmini or id in userListRotat:
            rawPostDate = data.find_element_by_class_name("post__time").get_attribute("datetime").split('T')[0]
            rawPostTime = data.find_element_by_class_name("post__time").get_attribute("datetime").split('T')[1].split('.')[0]
            postTimeStr = rawPostDate + ' ' + rawPostTime
            postTimeUtc = datetime.strptime(postTimeStr, "%Y-%m-%d %H:%M:%S")
            timeStamp = calendar.timegm(postTimeUtc.utctimetuple())
            localPostTime = datetime.fromtimestamp(timeStamp)
            localPostTimeStr = datetime.fromtimestamp(timeStamp).strftime('%Y-%m-%d %H:%M:%S')

            if localPostTime >= inputTime:
                rawRecord_dict = {
                    "ID": "",
                    "date": "",
                    "shift": "unknown",
                    "postTime":"",
                    "postMessage": "",
                    "status": "1"
                }
                rawRecord_dict['ID'] = id
                rawRecord_dict['postTime'] = localPostTimeStr.split(" ")[1]
                rawRecord_dict['date'] = localPostTimeStr.split(" ")[0]
                postMessage = data.find_element_by_class_name('post-message__text').text.strip()
                postMessageLower = postMessage.lower()
                rawRecord_dict['postMessage'] = postMessageLower
                if id in userListAdmini:
                    rawRecord_dict['shift'] = 'F'
                    if postMessageLower.__contains__('in') or postMessageLower.__contains__('out'):
                        rawRecord_dict['status'] = '0'
                if id in userListRotat:
                    shiftChk = int(localPostTimeStr.split(" ")[1].split(':')[0])
                    if postMessageLower.__contains__('in'):
                        if 21 <= shiftChk < 24:
                            rawRecord_dict['shift'] = 'D'
                        if 5 <= shiftChk < 8:
                            rawRecord_dict['shift'] = 'A'
                        if 9 <= shiftChk < 12:
                            rawRecord_dict['shift'] = 'B'
                        if 13 <= shiftChk < 16:
                            rawRecord_dict['shift'] = 'C'
                        rawRecord_dict['status'] = '0'
                    if postMessageLower.__contains__('out'):
                        if shiftChk < 12:
                            rawRecord_dict['shift'] = 'D'
                        rawRecord_dict['status'] = '0'
                if len(postMessageLower) > 10:
                    rawRecord_dict['status'] = '1'

                print(rawRecord_dict)
                rawRecordList.append(rawRecord_dict)

            else:
                continue
    #清洗資料
    dateSet = set()
    shiftList = ['A', 'B', 'C', 'D']
    for rawRecord in rawRecordList:
        dateSet.add(rawRecord['date'])
    for date in dateSet:
        starList = []
        startA = datetime.strptime(date + ' ' + '07:00:00', "%Y-%m-%d %H:%M:%S")
        starList.append(startA)
        startB = datetime.strptime(date + ' ' + '11:00:00', "%Y-%m-%d %H:%M:%S")
        starList.append(startB)
        startC = datetime.strptime(date + ' ' + '15:00:00', "%Y-%m-%d %H:%M:%S")
        starList.append(startC)
        startD = datetime.strptime(date + ' ' + '23:00:00', "%Y-%m-%d %H:%M:%S")
        starList.append(startD)
        endList = []
        endA = datetime.strptime(date + ' ' + '15:30:00', "%Y-%m-%d %H:%M:%S")
        endList.append(endA)
        endB = datetime.strptime(date + ' ' + '19:30:00', "%Y-%m-%d %H:%M:%S")
        endList.append(endB)
        endC = datetime.strptime(date + ' ' + '23:30:00', "%Y-%m-%d %H:%M:%S")
        endList.append(endC)
        endD = datetime.strptime(date + ' ' + '07:30:00', "%Y-%m-%d %H:%M:%S") + dt.timedelta(days=1)
        endList.append(endD)
        tmp0Record = []
        tmpIdSet = set()
        for record in rawRecordList:
            if record['date'] == date and record['status'] == '0' and record['shift'] != 'D':
                tmp0Record.append(record)
                tmpIdSet.add(record['ID'])
            if record['status'] == '0' and record['shift'] == 'D' and record['date'] == date \
                    and record['postMessage'].__contains__('in'):
                tmp0Record.append(record)
                tmpIdSet.add(record['ID'])
            if datetime.strptime(record['date'], "%Y-%m-%d") - datetime.strptime(date, "%Y-%m-%d") == dt.timedelta(days=1) \
                    and record['status'] == '0' and record['shift'] == 'D' and record['postMessage'].__contains__('out'):
                tmp0Record.append(record)
                tmpIdSet.add(record['ID'])
        for name in tmpIdSet:
            tmpRecord = []
            for record in tmp0Record:
                if record['ID'] == name:
                    tmpRecord.append(record)

            num = len(tmpRecord)
            attendanceStatus_dict = {
                'ID': '',
                'date': '',
                'shift': '',
                'signIn': '',
                'signOut': '',
                'workHours': '',
                'status': '',
            }
            if num >= 2 and tmpRecord[0]['postMessage'].__contains__('in') and tmpRecord[num-1]['postMessage'].__contains__('out'):
                attendanceStatus_dict['ID'] = name
                attendanceStatus_dict['date'] = date
                attendanceStatus_dict['shift'] = tmpRecord[0]['shift']
                attendanceStatus_dict['signIn'] = tmpRecord[0]['postTime']
                attendanceStatus_dict['signOut'] = tmpRecord[num-1]['postTime']
                signInTimeHourCal = tmpRecord[0]['postTime'].split(':')[0]
                signInTimeMinCal = tmpRecord[0]['postTime'].split(':')[1]
                signInTimeCal = date + ' ' + signInTimeHourCal + ':' + signInTimeMinCal + ':00'
                signOutTimeHourCal = tmpRecord[num-1]['postTime'].split(':')[0]
                signOutTimeMinCal = tmpRecord[num-1]['postTime'].split(':')[1]
                signOutTimeCal = tmpRecord[num-1]['date'] + ' ' + signOutTimeHourCal + ':' + signOutTimeMinCal + ':00'
                signInTime = datetime.strptime(signInTimeCal, "%Y-%m-%d %H:%M:%S")
                signOutTime = datetime.strptime(signOutTimeCal, "%Y-%m-%d %H:%M:%S")
                if tmpRecord[0]['shift'] == 'F':
                    if date not in holidayList:
                        if signOutTime - signInTime >= dt.timedelta(hours=9):
                            attendanceStatus_dict['status'] = '0'
                            attendanceStatus_dict['workHours'] = "+" + str(
                                signOutTime - signInTime - dt.timedelta(hours=9))
                        else:
                            attendanceStatus_dict['status'] = '3'
                            attendanceStatus_dict['workHours'] = "-" + str(
                                dt.timedelta(hours=9) - (signOutTime - signInTime))
                    if date in holidayList:
                        attendanceStatus_dict['status'] = '0'
                        attendanceStatus_dict['workHours'] = "+" + str(signOutTime - signInTime)
                else:
                    shiftNum = 0
                    for shift in shiftList:
                        if tmpRecord[0]['shift'] == shift:
                            if signOutTime - starList[shiftNum] >= dt.timedelta(hours=8.5):
                                attendanceStatus_dict['status'] = '0'
                                attendanceStatus_dict['workHours'] = "+" + str(signOutTime - starList[shiftNum] - dt.timedelta(hours=8.5))
                            else:
                                attendanceStatus_dict['status'] = '3'
                                attendanceStatus_dict['workHours'] = "-" + str(dt.timedelta(hours=8.5) - (signOutTime - starList[shiftNum]))
                            if signInTime > starList[shiftNum]:
                                if attendanceStatus_dict['status'] == '0':
                                    attendanceStatus_dict['status'] = '6'
                                else:
                                    attendanceStatus_dict['status'] = attendanceStatus_dict['status'] + ',6'
                            if signOutTime < endList[shiftNum]:
                                if attendanceStatus_dict['status'] == '0':
                                    attendanceStatus_dict['status'] = '7'
                                else:
                                    attendanceStatus_dict['status'] = attendanceStatus_dict['status'] + ',7'
                            break
                        shiftNum += 1
                if tmpRecord[0]['shift'] == 'unknown':
                    attendanceStatus_dict['status'] = '8'
                    attendanceStatus_dict['workHours'] = '-'
                if num > 2:
                    if attendanceStatus_dict['status'] == '0':
                        attendanceStatus_dict['status'] = '2'
                    else:
                        attendanceStatus_dict['status'] = attendanceStatus_dict['status'] + ',2'
                itDeskAttendanceList.append(attendanceStatus_dict)
            if num == 1:  # 將4錯誤訊息輸出
                tmpRecord[0]['status'] = '4'
                itDeskAttendanceList.append(tmpRecord[0])
                attendanceStatus_dict['ID'] = name
                attendanceStatus_dict['date'] = date
                attendanceStatus_dict['shift'] = tmpRecord[0]['shift']
                attendanceStatus_dict['status'] = '4'
                attendanceStatus_dict['workHours'] = '-'

                if tmpRecord[0]['postMessage'].__contains__('in'):
                    attendanceStatus_dict['signIn'] = tmpRecord[0]['postTime']
                    attendanceStatus_dict['signOut'] = '-'
                else:
                    attendanceStatus_dict['signIn'] = '-'
                    attendanceStatus_dict['signOut'] = tmpRecord[0]['postTime']

                itDeskAttendanceList.append(attendanceStatus_dict)
        if date not in holidayList:   # 將5錯誤訊息輸出
            for idChk in userListAdmini:
                if idChk not in tmpIdSet:
                    absence_dict = {
                        "ID": "",
                        "date": "",
                        "shift": "F",
                        "postTime": "-",
                        "postMessage": "-",
                        "status": "5"
                    }
                    absence_dict['ID'] = idChk
                    absence_dict['date'] = date
                    itDeskAttendanceList.append(absence_dict)
    for i in range(0, len(rawRecordList), 1):   # 將1錯誤訊息輸出
        if rawRecordList[i]['status'] == '1':
            itDeskAttendanceList.append(rawRecordList[i])
    for i in range(0, len(itDeskAttendanceList), 1):    # 將2,3,6,7,8錯誤訊息輸出
        statusList = ['0', '1', '4', '5']
        if itDeskAttendanceList[i]['status'] not in statusList:
            workHoursChk_dict = {
                "ID": "",
                "date": "",
                "shift": "",
                "postTime": "-",
                "postMessage": "-",
                "status": ""
            }
            workHoursChk_dict['ID'] = itDeskAttendanceList[i]['ID']
            workHoursChk_dict['date'] = itDeskAttendanceList[i]['date']
            workHoursChk_dict['shift'] = itDeskAttendanceList[i]['shift']
            workHoursChk_dict['status'] = itDeskAttendanceList[i]['status']
            itDeskAttendanceList.append(workHoursChk_dict)

    # 寫成Excel(日期分頁)
    attendanceXlsx = Workbook()
    sheet = attendanceXlsx.active
    attendanceXlsx.remove_sheet(sheet)

    rawRecordSheet = attendanceXlsx.create_sheet('rawRecord')
    rawRecordSheet.cell(row=1, column=1, value='date')
    rawRecordSheet.cell(row=1, column=2, value='ID')
    rawRecordSheet.cell(row=1, column=3, value='postTime')
    rawRecordSheet.cell(row=1, column=4, value='postMessage')
    rawRecordSheet.column_dimensions['A'].width = 15
    rawRecordSheet.column_dimensions['B'].width = 15
    rawRecordSheet.column_dimensions['C'].width = 15
    rawRecordSheet.column_dimensions['D'].width = 15
    rowNum = 2
    for raw in rawRecordList:
        rawRecordSheet.cell(row=rowNum, column=1, value=raw['date'])
        rawRecordSheet.cell(row=rowNum, column=2, value=raw['ID'])
        rawRecordSheet.cell(row=rowNum, column=3, value=raw['postTime'])
        rawRecordSheet.cell(row=rowNum, column=4, value=raw['postMessage'])
        rowNum += 1

    statusReportSheet = attendanceXlsx.create_sheet('statusReport')
    statusReportSheet.cell(row=1, column=1, value='status code')
    statusReportSheet.cell(row=1, column=2, value='description')
    statusReportSheet.cell(row=2, column=1, value='1')
    statusReportSheet.cell(row=2, column=2, value='incorrect postMessage records')
    statusReportSheet.cell(row=3, column=1, value='2')
    statusReportSheet.cell(row=3, column=2, value='more than 2 postMessage records in a day')
    statusReportSheet.cell(row=4, column=1, value='3')
    statusReportSheet.cell(row=4, column=2, value='Work less than 8.5 hours')
    statusReportSheet.cell(row=5, column=1, value='4')
    statusReportSheet.cell(row=5, column=2, value='only 1 postMessage record in a day')
    statusReportSheet.cell(row=6, column=1, value='5')
    statusReportSheet.cell(row=6, column=2, value='Absence')
    statusReportSheet.cell(row=7, column=1, value='6')
    statusReportSheet.cell(row=7, column=2, value='Late')
    statusReportSheet.cell(row=8, column=1, value='7')
    statusReportSheet.cell(row=8, column=2, value='Early leave')
    statusReportSheet.cell(row=9, column=1, value='8')
    statusReportSheet.cell(row=9, column=2, value='Unknown Shift')

    statusReportSheet.cell(row=11, column=1, value='postDate')
    statusReportSheet.cell(row=11, column=2, value='shift')
    statusReportSheet.cell(row=11, column=3, value='ID')
    statusReportSheet.cell(row=11, column=4, value='status code')
    statusReportSheet.cell(row=11, column=5, value='postTime')
    statusReportSheet.cell(row=11, column=6, value='postMessage')
    statusReportSheet.cell(row=1, column=6, value='Shift')
    statusReportSheet.cell(row=2, column=6, value='A: 07:00-15:30')
    statusReportSheet.cell(row=3, column=6, value='B: 11:00-19:30')
    statusReportSheet.cell(row=4, column=6, value='C: 15:00-23:30')
    statusReportSheet.cell(row=5, column=6, value='D: 23:00-07:30')
    statusReportSheet.cell(row=6, column=6, value='F: 09:00-18:00')
    statusReportSheet.column_dimensions['A'].width = 15
    statusReportSheet.column_dimensions['B'].width = 15
    statusReportSheet.column_dimensions['C'].width = 15
    statusReportSheet.column_dimensions['D'].width = 15
    statusReportSheet.column_dimensions['E'].width = 15
    statusReportSheet.column_dimensions['F'].width = 15


    R = 12
    dateList = list(dateSet)
    list.sort(dateList)
    for date in dateList:
        sheet = attendanceXlsx.create_sheet(date)
        sheet['A1'] = 'Date'
        sheet['B1'] = 'Shift'
        sheet['C1'] = 'ID'
        sheet['D1'] = 'Sign In'
        sheet['E1'] = 'Sign Out'
        sheet['F1'] = 'Work Hours'
        sheet['G1'] = 'Status'
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 15

        r = 2
        for i in range(0, len(itDeskAttendanceList), 1):
            if itDeskAttendanceList[i]['date'] == date and len(itDeskAttendanceList[i]) == 7:
                sheet.cell(row=r, column=1, value=date)
                sheet.cell(row=r, column=2, value=itDeskAttendanceList[i]['shift'])
                sheet.cell(row=r, column=3, value=itDeskAttendanceList[i]['ID'])
                sheet.cell(row=r, column=4, value=itDeskAttendanceList[i]['signIn'])
                sheet.cell(row=r, column=5, value=itDeskAttendanceList[i]['signOut'])
                sheet.cell(row=r, column=6, value=itDeskAttendanceList[i]['workHours'])
                sheet.cell(row=r, column=7, value=itDeskAttendanceList[i]['status'])
                r += 1
            if itDeskAttendanceList[i]['date'] == date and len(itDeskAttendanceList[i]) == 6:
                statusReportSheet.cell(row=R, column=1, value=itDeskAttendanceList[i]['date'])
                statusReportSheet.cell(row=R, column=2, value=itDeskAttendanceList[i]['shift'])
                statusReportSheet.cell(row=R, column=3, value=itDeskAttendanceList[i]['ID'])
                statusReportSheet.cell(row=R, column=4, value=itDeskAttendanceList[i]['status'])
                statusReportSheet.cell(row=R, column=5, value=itDeskAttendanceList[i]['postTime'])
                statusReportSheet.cell(row=R, column=6, value=itDeskAttendanceList[i]['postMessage'])
                R += 1

    attendanceXlsx.save('/Users/yunhan/Desktop/itDesk_Attendance.xlsx')
    print('Done')
    web.quit()

    #寫成jason
    # encodedjson = json.dumps(itDeskAttendanceList, ensure_ascii=False)
    # with open("/Users/yunhan/Desktop/itAttendance.json", "w") as f:
    #     f.write(encodedjson)
    #     f.close()
    # web.quit()
